from __future__ import annotations

import hashlib
import io
import ipaddress
import socket
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from django.conf import settings

MAX_DOCUMENT_BYTES = int(getattr(settings, "FORGEAI_MAX_DOCUMENT_BYTES", 25 * 1024 * 1024))
CHUNK_SIZE = 3500
CHUNK_OVERLAP = 350

class DocumentIngestionError(RuntimeError):
    pass

def _validate_public_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise DocumentIngestionError("Only public HTTP(S) document URLs are supported.")
    try:
        addresses = {row[4][0] for row in socket.getaddrinfo(parsed.hostname, parsed.port or 443)}
    except socket.gaierror as exc:
        raise DocumentIngestionError("The document host could not be resolved.") from exc
    for address in addresses:
        ip = ipaddress.ip_address(address)
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast:
            raise DocumentIngestionError("Private or local network document URLs are blocked.")

def download_document(url: str) -> tuple[bytes, str]:
    _validate_public_url(url)
    try:
        with requests.get(url, stream=True, timeout=35, allow_redirects=True, headers={"User-Agent": "ForgeGov/2.2"}) as response:
            response.raise_for_status()
            final_url = response.url
            _validate_public_url(final_url)
            content_type = response.headers.get("content-type", "").split(";", 1)[0].strip().lower()
            chunks: list[bytes] = []
            total = 0
            for chunk in response.iter_content(65536):
                if not chunk:
                    continue
                total += len(chunk)
                if total > MAX_DOCUMENT_BYTES:
                    raise DocumentIngestionError("The document exceeds the 25 MB ingestion limit.")
                chunks.append(chunk)
            return b"".join(chunks), content_type
    except requests.RequestException as exc:
        raise DocumentIngestionError("ForgeGov could not download the source document.") from exc

def _clean(text: str) -> str:
    return "\n".join(line.strip() for line in str(text or "").replace("\x00", " ").splitlines() if line.strip())

def extract_document(data: bytes, filename: str, content_type: str = "") -> list[tuple[int | None, str | None, str]]:
    suffix = Path(filename.lower()).suffix
    if suffix == ".pdf" or content_type == "application/pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        return [(index, None, _clean(page.extract_text() or "")) for index, page in enumerate(reader.pages, 1) if _clean(page.extract_text() or "")]
    if suffix == ".docx" or content_type.endswith("wordprocessingml.document"):
        from docx import Document
        document = Document(io.BytesIO(data))
        text = _clean("\n".join(paragraph.text for paragraph in document.paragraphs))
        return [(None, None, text)] if text else []
    if suffix in {".xlsx", ".xlsm"} or "spreadsheetml" in content_type:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sections = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
            text = _clean("\n".join(rows))
            if text:
                sections.append((None, sheet.title, text))
        return sections
    text = data.decode("utf-8", errors="replace")
    if suffix in {".html", ".htm"} or content_type == "text/html":
        text = BeautifulSoup(text, "html.parser").get_text("\n")
    text = _clean(text)
    return [(None, None, text)] if text else []

def chunk_sections(sections: list[tuple[int | None, str | None, str]]):
    for page_number, section, text in sections:
        start = 0
        ordinal = 0
        while start < len(text):
            end = min(len(text), start + CHUNK_SIZE)
            piece = text[start:end].strip()
            if piece:
                ordinal += 1
                yield page_number, section, ordinal, piece
            if end == len(text):
                break
            start = max(start + 1, end - CHUNK_OVERLAP)

def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()
